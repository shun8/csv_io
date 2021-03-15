#!/usr/bin/env python3
import os
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import re
import sys
import yaml

import postgresclient

# 引数判定
if len(sys.argv) != 4:
    print("Usage: python " + os.path.basename(__file__) + " <format_file> <output_file_path> <yyyymm>")
    sys.exit(1)

p_format_file = sys.argv[1]
p_output_file_path = sys.argv[2]
p_yyyymm = sys.argv[3]

# NULL行名
null_col_name = "VALUE_IS_NULL"

# sqlファイルに変数を渡す際の変数名の表記ゆれに対応 (変数名なしの%s指定には未対応)
yyyymm_val_names = ["ym", "yyyymm"]
yyyymm_val_dict = {x: p_yyyymm for x in yyyymm_val_names}

# フォーマット定義読み込み
with open(p_format_file, "r", encoding="utf-8") as format_file:
    format_config = yaml.safe_load(format_file)

# DB接続情報
l_dirname = os.path.dirname(os.path.abspath(__file__))
with open(os.path.normpath(os.path.join(l_dirname, "../db_connection.yaml")), "r") as db_config:
    l_db_config = yaml.safe_load(db_config)
    db_client = postgresclient.PostgresClient(l_db_config["postgres"])
# SQLファイル配置場所
sql_dir = os.path.normpath(os.path.join(l_dirname, "../sql"))

# 数字からエクセル列アルファベットへの変換関数（1→A、27→AAみたいな変換）
# https://www.engineer-log.com/entry/2018/03/15/openpyxl-frame-border
def to_alpha(num):
    h=int((num-1-26)/(26*26))
    i=int((num-1-(h*26*26))/26)
    j=int(num-(i*26)-(h*26*26))
    alpha=''
    for k in h,i,j:
        if k!=0:
            alpha+=chr(k+64)
    return alpha

# アドレス計算用関数
def calc_header_span(p_index, p_col_headers):
    if p_index >= len(p_col_headers) - 1:
        return 1
    return p_col_headers[p_index+1]["size"] * calc_header_span(p_index+1, p_col_headers)

# ヘッダ描画用関数 必要なら引数にスタイル追加
def draw_headers(ws, p_row_i, p_start_col_i, p_header_i, p_col_headers, p_col_header_confs):
    l_row_span = p_col_header_confs[p_header_i].get("row_span")
    l_row_offset = p_col_header_confs[p_header_i].get("row_offset")
    if l_row_span is None:
        l_row_span = 1
    if l_row_offset is None:
        l_row_offset = 0

    l_col_i = p_start_col_i
    for h_txt in col_headers[p_header_i]["indexes"].keys():
        # value
        ws.cell(p_row_i + l_row_offset , l_col_i, value=h_txt)
        # styles
        for i in range(p_row_i, p_row_i + l_row_span):
            apply_styles(ws, i, l_col_i, p_col_header_confs[p_header_i]["style"])

        # child headers
        if p_header_i < len(p_col_headers) - 1:
            draw_headers(ws, p_row_i + l_row_span, l_col_i,
                    p_header_i + 1, p_col_headers, p_col_header_confs)

        l_col_span = col_headers[p_header_i]["span"]
        l_col_i += l_col_span
        # merge
        if p_col_header_confs[p_header_i].get("merge"):
            ws.merge_cells(
                    start_row=p_row_i, start_column=l_col_i - l_col_span,
                    end_row=p_row_i, end_column=l_col_i - 1)

def calc_col_header_row_span(p_col_header_confs):
    l_count = len(p_col_header_confs)
    for conf_i in p_col_header_confs:
        l_row_span = conf_i.get("row_span")
        if l_row_span:
            l_count += l_row_span - 1
    return l_count

def apply_styles(ws, p_row_i, p_col_i, p_style):
    if p_style is None:
        return

    if p_style.get("width"):
        ws.column_dimensions[to_alpha(p_col_i)].width = p_style.get("width")
    if p_style.get("font"):
        ws.cell(p_row_i, p_col_i).font = Font(
                name=p_style["font"]["name"], size=p_style["font"]["size"],
                bold=p_style["font"]["bold"], color=p_style["font"]["color"])
    if p_style.get("fill"):
        ws.cell(p_row_i, p_col_i).fill = PatternFill(
                patternType=p_style["fill"]["pattern_type"],
                fgColor=p_style["fill"]["fg_color"])
    if p_style.get("alignment"):
        ws.cell(p_row_i, p_col_i).alignment = Alignment(
                horizontal=p_style["alignment"]["holizontal"],
                vertical=p_style["alignment"]["vertical"])
    if p_style.get("border"):
        l_border = Border(
                    right=ws.cell(p_row_i, p_col_i).border.right,
                    left=ws.cell(p_row_i, p_col_i).border.left,
                    top=ws.cell(p_row_i, p_col_i).border.top,
                    bottom=ws.cell(p_row_i, p_col_i).border.bottom)
        for side_i in ["top", "bottom", "left", "right"]:
            if p_style["border"].get(side_i):
                setattr(l_border, side_i, Side(
                        border_style=p_style["border"][side_i]["border_style"],
                        color=p_style["border"][side_i]["color"]))
        ws.cell(p_row_i, p_col_i).border = l_border

# ---- 出力ファイル作成処理 ----
# フォーマット確認(今はxlsxしかないので実質意味なし)
if format_config.get("format") not in ["xlsx"]:
    print(format_config.get("format") + " is not valid format.")
    sys.exit(1)

# デフォルトスタイル取得
default_style = format_config.get("style")

# -- xlsx出力処理 --
# ベースになるファイル読み込み
basefile = format_config.get("basefile")
if basefile is None:
    wb = openpyxl.Workbook()
elif os.path.isfile(basefile):
    wb = openpyxl.load_workbook(p_output_file_path)
else:
    print("File could not be found " + basefile + ".")
    sys.exit(1)

# create sheet
for sheet in format_config["sheets"]:
    if sheet["name"] not in wb.sheetnames:
        wb.create_sheet(index=sheet["index"], title=sheet["name"])
    ws = wb[sheet["name"]]

    row_i = sheet["row_padding"] + 1
    col_i = sheet["col_padding"] + 1
    row_header_span = sheet["row_header_span"] 

    # query column headers
    col_headers = []
    col_header_confs = sorted(sheet["col_headers"], key=lambda x: x["index"])
    for col_header_conf in col_header_confs:
        ds_conf = col_header_conf["source"]
        with open(os.path.normpath(os.path.join(sql_dir, ds_conf["sql"])), "r") as l_sqlfile:
            l_sql = l_sqlfile.read()
            # :始まりの変数をpsycopg2で入力できる形式に変換
            l_sql = re.sub(':([A-Za-z-_]+)', '%(\\1)s', l_sql)
        l_table = db_client.execute(l_sql, yyyymm_val_dict)
        l_table = sorted(l_table, key=lambda x: x[ds_conf["order"]])
        l_col_h = {
            "size": len(l_table),
            "indexes": {x[ds_conf["data"]]: x[ds_conf["order"]] for x in l_table}
        }
        col_headers.append(l_col_h)
    for i in range(len(col_headers)):
        col_headers[i]["span"] = calc_header_span(i, col_headers)

    # draw title
    for h_i, col_header_conf in enumerate(col_header_confs):
        l_text = col_header_conf["header_title"]["text"]
        try:
            l_text = l_text.format(yyyy=p_yyyymm[:4], mm=p_yyyymm[4:6])
        except IndexError:
            pass
        ws.cell(row_i + h_i, col_i, value=l_text)
        apply_styles(ws, row_i + h_i, col_i, col_header_conf["header_title"]["style"])

        # merge
        l_row_span = col_header_conf.get("row_span")
        if col_header_conf["header_title"].get("merge") and l_row_span:
            ws.merge_cells(
                    start_row=row_i + h_i, start_column=col_i,
                    end_row=row_i + h_i + l_row_span - 1, end_column=col_i)

    # draw column headers
    draw_headers(ws, row_i, col_i + row_header_span, 0, col_headers, col_header_confs)
    row_i += calc_col_header_row_span(col_header_confs)

    if sheet.get("freeze_panes"):
        ws.freeze_panes = to_alpha(row_header_span+1) + str(row_i)

    # query bodies
    bodies = []
    body_confs = sorted(sheet["bodies"], key=lambda x: x["index"])
    for body_conf in body_confs:
        ds_conf = body_conf["source"]
        rh_ds_conf = ds_conf["group"]["row_header"]["source"]
        with open(os.path.normpath(os.path.join(sql_dir, rh_ds_conf["sql"])), "r") as l_sqlfile:
            l_sql = l_sqlfile.read()
            l_sql = re.sub(':([A-Za-z-_]+)', '%(\\1)s', l_sql)
        l_rh_table = db_client.execute(l_sql, yyyymm_val_dict)
        l_rh_table = sorted(l_rh_table, key=lambda x: x[rh_ds_conf["order"]])
        l_rh_indexes = {x[rh_ds_conf["data"]]: x[rh_ds_conf["order"]] for x in l_rh_table}
        l_rh_indexes[null_col_name] = len(l_rh_table)

        with open(os.path.normpath(os.path.join(sql_dir, ds_conf["sql"])), "r") as l_sqlfile:
            l_sql = l_sqlfile.read()
            l_sql = re.sub(':([A-Za-z-_]+)', '%(\\1)s', l_sql)
        l_table = db_client.execute(l_sql, yyyymm_val_dict)
        l_ch_columns = ds_conf["group"]["col_headers"]
        l_ch_columns = sorted(l_ch_columns, key=lambda x: x["header_index"])
        l_body = {
            "table": l_table,
            "rh_indexes": l_rh_indexes,
            "rh_column_name": ds_conf["group"]["row_header"]["column_name"],
            "ch_column_names": [x["column_name"] for x in l_ch_columns]
        }
        bodies.append(l_body)

    # draw bodies
    for body, body_conf in zip(bodies, body_confs):
        is_null_error = False
        l_start_row = row_i
        for l_row in body["table"]:
            l_row_i = row_i + int(body["rh_indexes"][l_row[body["rh_column_name"]]])
            if [l_row[body["rh_column_name"]]] == null_col_name:
                is_null_error = True
            l_col_i = col_i + row_header_span
            for i in range(len(col_headers)):
                l_h_idx = int(col_headers[i]["indexes"][l_row[body["ch_column_names"][i]]])
                l_col_i += l_h_idx * col_headers[i]["span"]
            ws.cell(l_row_i, l_col_i, value=l_row[ds_conf["data"]])
            apply_styles(ws, l_row_i, l_col_i, default_style)
            ws.cell(l_row_i, l_col_i).number_format = default_style["number_format"]
            # ds_conf["cell_style"]
        # draw row headers
        if not is_null_error:
            body["rh_indexes"].pop(null_col_name)
        for h_txt in body["rh_indexes"].keys():
            ws.cell(row_i, col_i, value=h_txt)
            l_style = body_conf["row_header_style"] if body_conf.get("row_header_style") else default_style
            apply_styles(ws, row_i, col_i, l_style)
            row_i += 1

        # fill by 0
        for i in range(l_start_row, row_i):
            for j in range(col_headers[0]["size"] * col_headers[0]["span"]):
                if ws.cell(i, col_i + row_header_span + j).value is None:
                    ws.cell(i, col_i + row_header_span + j).value = 0
                    apply_styles(ws, i, col_i + row_header_span + j, default_style)
        
        # last row border
        if body_conf.get("last_row_border"):
            for i in range(col_i + row_header_span, ws.max_column + 1):
                ws.cell(row_i - 1, i).border = Border(bottom=Side(
                    border_style=body_conf["last_row_border"]["border_style"],
                    color=body_conf["last_row_border"]["color"]))

    # column border by top header
    top_header_conf = col_header_confs[0]
    last_col_line = top_header_conf.get("last_col_border")
    if last_col_line is not None:
        for i in range(1, ws.max_row + 1):
            for j in range(len(col_headers[0]["indexes"])):
                l_j = row_header_span + (j + 1) * col_headers[0]["span"]
                ws.cell(i, l_j).border = Border(
                    right=Side(
                        border_style=top_header_conf["last_col_border"]["border_style"],
                        color=top_header_conf["last_col_border"]["color"]),
                    left=ws.cell(i, l_j).border.left,
                    top=ws.cell(i, l_j).border.top,
                    bottom=ws.cell(i, l_j).border.bottom)

wb.save(p_output_file_path)
#Completed.
print("Successfully exported.")
sys.exit(0)
