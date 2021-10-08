#!/usr/bin/env python3
import os
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import logging.config
import re
import sys
import yaml

base = os.path.dirname(os.path.abspath(__file__))
LOGGING_CONF_FILE = os.path.normpath(os.path.join(base, "../logging.ini"))

# loggerを初期化
if os.path.exists(LOGGING_CONF_FILE):
    # logging.iniファイルがある場合設定を読み込む
    logging.config.fileConfig(LOGGING_CONF_FILE)
    logger = logging.getLogger("regular")
else:
    # logging.iniファイルがない場合すべてコンソールに出力
    handler = logging.StreamHandler()
    handler.setLevel(logging.DEBUG)

    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)
    logger.addHandler(handler)

# logging
def logging(func):
    def wrapper(*args, **kwargs):
        try:
            logger.info("Start: " + func.__name__)
            logger.info("Args: " + str(args) + str(kwargs))
            retval = func(*args, **kwargs)
            logger.info("End: " + func.__name__)
            return retval
        except:
            exc_info = sys.exc_info()
            logger.error("Error: %s,%s" % (exc_info[0], exc_info[1]))
            raise
    return wrapper

class XLSXGenerator:
    def __init__(self, db_client, sql_dir):
        # DB接続クライアント
        self._db_client = db_client
        # SQLファイル配置場所
        self._sql_dir = sql_dir

    @logging
    def gen_xlsx(self, format_file, yyyymm):
        # read format config
        with open(format_file, "r", encoding="utf-8") as l_file:
            format_config = yaml.safe_load(l_file)
        # create workbook
        basefile = format_config.get("basefile")
        if basefile is not None and not os.path.isfile(basefile):
            logger.error("File could not be opened " + basefile + ".")
            sys.exit(1)
        wb = self._create_workbook(basefile)

        # create worksheets
        for sheet_conf in format_config["sheets"]:
            if sheet_conf["name"] not in wb.sheetnames:
                wb.create_sheet(index=sheet_conf["index"], title=sheet_conf["name"])
            ws = wb[sheet_conf["name"]]

            # create column headers
            col_headers = self._create_column_headers(sheet_conf["col_headers"], yyyymm)
            # Create Bodies
            bodies = self._create_bodies(sheet_conf["bodies"], col_headers, yyyymm)

            # draw column headers
            row_offset = sheet_conf["row_padding"]
            col_offset = sheet_conf["col_padding"]
            self._draw_header_titles(ws, col_headers, row_offset, col_offset, yyyymm)
            col_body_offset = col_offset + sheet_conf["row_header_span"]
            self._draw_column_headers(ws, col_headers, row_offset, col_body_offset)

            # draw row headers
            row_body_offset = row_offset + sum([x.row_span for x in col_headers])
            self._draw_row_headers(ws, bodies, row_body_offset, col_offset)
            # draw bodies
            default_style = format_config.get("style")
            self._draw_bodies(ws, bodies, row_body_offset, col_body_offset, default_style)
            # fill by 0
            self._fill_none_by_zero(ws, row_body_offset, col_body_offset)
            # freeze panes
            if sheet_conf.get("freeze_panes"):
                ws.freeze_panes = self._to_alpha(col_body_offset + 1) + str(row_body_offset + 1)

            # last row border by body
            self._apply_last_row_borders(ws, bodies, row_body_offset, col_offset)
            # last column border by top header span
            self._apply_last_column_borders(ws, col_headers[0], col_body_offset)

        # delete default worksheet
        if basefile is None:
            sheet_names = [x["name"] for x in format_config["sheets"]]
            unused_sheets = list(set(wb.get_sheet_names()) - set(sheet_names))
            for sheet_name in unused_sheets:
                ws = wb.get_sheet_by_name(sheet_name)
                wb.remove_sheet(ws)

        return wb

    @logging
    def gen_xlsx_like_csv(self, format_file, sql_params={}):
        # read format config
        with open(format_file, "r", encoding="utf-8") as l_file:
            format_config = yaml.safe_load(l_file)
        # create workbook
        basefile = format_config.get("basefile")
        if basefile is not None and not os.path.isfile(basefile):
            logger.error("File could not be opened " + basefile + ".")
            sys.exit(1)
        wb = self._create_workbook(basefile)

        additional_sheet_names = []
        # create worksheets
        for sheet_conf in format_config["sheets"]:
            limit = sheet_conf.get("limit")
            offset = 0         
            sheet_count = 1       
            while True:
                if sheet_count == 1:
                    sheet_name = sheet_conf["name"]
                else:
                    sheet_name = sheet_conf["name"] + "_" + str(sheet_count)

                if sheet_name not in wb.sheetnames:
                    if sheet_count == 1:
                        wb.create_sheet(index=sheet_conf["index"], title=sheet_name)
                        ws = wb[sheet_name]
                    else:
                        ws = wb.copy_worksheet(wb[sheet_conf["name"]])
                        ws.title = sheet_name
                        # clear values
                        row_offset = sheet_conf["row_padding"]
                        for row in ws.iter_rows(min_row=row_offset+1):
                            for cell in row:
                                cell.value = None
                else:
                    ws = wb[sheet_name]

                sql_params.update({"limit": limit, "offset": offset})
                result = self._query_by_params(sheet_conf["source"]["sql"], sql_params)
                if not result: break

                row_offset = sheet_conf["row_padding"]
                col_offset = sheet_conf["col_padding"]
                default_style = format_config.get("style")

                # draw header
                if sheet_conf["header"]:
                    if sheet_conf.get("header_style"):
                        header_style = sheet_conf.get("header_style")
                    else:
                        header_style = default_style
                    row_index = row_offset + 1
                    XLSXGenerator._draw_row(
                            ws, result[0].keys(), row_index, col_offset, header_style)
                    row_offset = row_offset + 1

                # draw body
                if sheet_conf.get("style"):
                    body_style = sheet_conf.get("style")
                else:
                    body_style = default_style
                XLSXGenerator._draw_rows(ws, result, row_offset, col_offset, body_style)

                if limit is not None:
                    offset = offset + limit
                else:
                    break

        # delete default worksheet
        if basefile is None:
            sheet_names = [x["name"] for x in format_config["sheets"]]
            unused_sheets = list(set(wb.get_sheet_names())
                    - set(sheet_names) - set(additional_sheet_names))
            for sheet_name in unused_sheets:
                ws = wb.get_sheet_by_name(sheet_name)
                wb.remove_sheet(ws)

        return wb

    def _query_header(self, data_source_conf, yyyymm=None):
        table = self._query_by_yyyymm(data_source_conf["sql"], yyyymm)
        return sorted(table, key=lambda x: x[data_source_conf["order"]])

    def _query_by_yyyymm(self, sqlfile_name, yyyymm):
        # ym指定する際の変数名リスト
        yyyymm_var_names = ["ym", "yyyymm"]
        yyyymm_var_dict = {x: yyyymm for x in yyyymm_var_names}
        return self._query_by_params(sqlfile_name, yyyymm_var_dict)

    def _query_by_params(self, sqlfile_name, sql_params={}):
        # sqlファイルに変数を渡す際の変数名の表記ゆれに対応 (変数名なしの%s指定には未対応)
        with open(os.path.normpath(os.path.join(self._sql_dir, sqlfile_name)), "r") as sqlfile:
            sql = sqlfile.read()
            # :始まりの変数をpsycopg2で入力できる形式に変換
            sql = re.sub(':([A-Za-z-_]+)', '%(\\1)s', sql)
        return self._db_client.execute(sql, sql_params)

    def _create_column_headers(self, headers_conf, yyyymm=None):
        col_headers = []
        headers_conf_sorted = sorted(headers_conf, key=lambda x: x["index"])
        for header_conf in headers_conf_sorted:    
            ds_conf = header_conf["source"]
            table = self._query_header(ds_conf, yyyymm)
            col_headers.append(ColumnHeader(header_conf, table))
        self._set_headers_span(col_headers)
        return col_headers

    def _create_row_header(self, header_conf, yyyymm=None):
        ds_conf = header_conf["source"]
        table = self._query_header(ds_conf, yyyymm)
        return RowHeader(header_conf, table)

    def _create_bodies(self, bodies_conf, col_headers, yyyymm=None):
        bodies = []
        bodies_conf_sorted = sorted(bodies_conf, key=lambda x: x["index"])
        for body_conf in bodies_conf_sorted:
            rh_conf = body_conf["row_header"]    
            rh_table = self._query_header(rh_conf["source"], yyyymm)
            body_table = self._query_by_yyyymm(body_conf["source"]["sql"], yyyymm)
            bodies.append(Body(body_conf, body_table,
                    RowHeader(rh_conf, rh_table), col_headers))
        return bodies

    @staticmethod
    def _create_workbook(basefile=None):
        if basefile is None:
            return openpyxl.Workbook()
        elif os.path.isfile(basefile):
            return openpyxl.load_workbook(basefile)

    @staticmethod
    def _draw_rows(ws, rows, row_offset, col_offset, style=None):
        row_i = row_offset + 1
        for row in rows:
            XLSXGenerator._draw_row(ws, row.values(), row_i, col_offset, style)
            row_i = row_i + 1

    @staticmethod
    def _draw_row(ws, values, row_index, col_offset, style=None):
        col_i = col_offset + 1
        for val in values:
            ws.cell(row_index, col_i, value=val)
            XLSXGenerator._apply_cell_styles(ws, row_index, col_i, style)
            col_i = col_i + 1

    @staticmethod
    def _draw_column_headers(ws, col_headers, row_offset, col_offset):
        times = 1
        row_index = row_offset + 1
        for col_header in col_headers:
            col_index = col_offset + 1
            for _ in range(times):
                for header_text in col_header.indexes.keys():
                    ws.cell(row_index + col_header.row_offset, col_index, value=header_text)
                    for i in range(row_index, row_index + col_header.row_span):
                        XLSXGenerator._apply_cell_styles(
                                ws, i, col_index, col_header.style_conf)
                    if col_header.merge:
                        # set value for merge
                        ws.cell(row_index, col_index, value=header_text)
                        ws.merge_cells(
                            start_row=row_index,
                            start_column=col_index,
                            end_row=row_index + col_header.row_span - 1,
                            end_column=col_index + col_header.span - 1)
                    col_index += col_header.span
            times *= col_header.count
            row_index += col_header.row_span

    @staticmethod
    def _draw_header_titles(ws, col_headers, row_offset, col_offset, yyyymm=None):
        row_index = row_offset + 1
        for col_header in col_headers:
            col_index = col_offset + 1
            text = col_header.header_title_conf.get("text")
            try:
                if yyyymm:                
                    text = text.format(yyyy=yyyymm[:4], mm=yyyymm[4:6])
            except IndexError:
                pass
            ws.cell(row_index + col_header.row_offset, col_index, value=text)
            for i in range(row_index, row_index + col_header.row_span):
                XLSXGenerator._apply_cell_styles(
                        ws, i, col_index, col_header.header_title_conf.get("style"))
            if col_header.header_title_conf.get("merge"):
                # set value for merge
                ws.cell(row_index, col_index, value=text)
                ws.merge_cells(
                    start_row=row_index,
                    start_column=col_index,
                    end_row=row_index + col_header.row_span - 1,
                    end_column=col_index)
            row_index += col_header.row_span

    @staticmethod
    def _draw_row_headers(ws, bodies, row_offset, col_offset):
        row_index = row_offset + 1
        col_index = col_offset + 1
        for body in bodies:
            for header_text in body.row_header.indexes.keys():
                ws.cell(row_index, col_index, value=header_text)
                XLSXGenerator._apply_cell_styles(
                        ws, row_index, col_index, body.row_header.style_conf)
                row_index += 1
            if body.has_null_error:
                ws.cell(row_index, col_index, value="ERROR: NO MASTER DATA")
                row_index += 1

    @staticmethod
    def _draw_bodies(ws, bodies, row_offset, col_offset, default_style=None):
        l_row_offset = row_offset
        for body in bodies:
            start_row = l_row_offset + 1
            for item in body.items:
                row_i = l_row_offset + item.row_index
                col_i = col_offset + item.col_index
                try:
                    ws.cell(row_i, col_i, value=int(item.value))
                except ValueError:
                    ws.cell(row_i, col_i, value=item.value)
            l_row_offset += body.row_header.count
            if body.has_null_error:
                l_row_offset += 1            
            for i in range(start_row, l_row_offset + 1):
                for j in range(col_offset, ws.max_column + 1):
                    XLSXGenerator._apply_cell_styles(ws, i, j, default_style)
                    XLSXGenerator._apply_cell_styles(ws, i, j, body.style_conf)

    @staticmethod
    def _fill_none_by_zero(ws, row_offset, col_offset):
        for i in range(row_offset + 1, ws.max_row + 1):
            for j in range(col_offset + 1, ws.max_column + 1):
                if ws.cell(i, j).value is None:
                    ws.cell(i, j).value = 0

    @staticmethod
    def _apply_last_row_borders(ws, bodies, row_offset, col_offset):
        row_index = row_offset
        for body in bodies:
            row_index += body.row_header.count
            if body.has_null_error:
                row_index += 1
            if body.last_row_border_conf:
                for j in range(col_offset + 1, ws.max_column + 1):
                    XLSXGenerator._set_border(
                            ws.cell(row_index, j), body.last_row_border_conf)

    @staticmethod
    def _apply_last_column_borders(ws, top_header, col_offset, row_offset=0):
        if top_header.last_col_border_conf:
            for j in range(col_offset + top_header.span, ws.max_column + 1, top_header.span):
                for i in range(row_offset + 1, ws.max_row + 1):
                    XLSXGenerator._set_border(
                            ws.cell(i, j), top_header.last_col_border_conf)

    @staticmethod
    def _apply_cell_styles(ws, row_i, col_i, style_conf):
        if style_conf is None:
            return

        if style_conf.get("width"):
            ws.column_dimensions[XLSXGenerator._to_alpha(col_i)].width = style_conf["width"]

        if style_conf.get("font"):
            ws.cell(row_i, col_i).font = Font(**style_conf["font"])
        if style_conf.get("fill"):
            ws.cell(row_i, col_i).fill = PatternFill(**style_conf["fill"])
        if style_conf.get("alignment"):
            ws.cell(row_i, col_i).alignment = Alignment(**style_conf["alignment"])
        if style_conf.get("border"):
            XLSXGenerator._set_border(ws.cell(row_i, col_i), style_conf["border"])
        
        if style_conf.get("number_format"):
            ws.cell(row_i, col_i).number_format = style_conf["number_format"]

    @staticmethod
    def _set_border(cell, border_conf):
        if border_conf.get("top"):
            top = Side(**border_conf["top"])
        else:
            top = cell.border.top
        if border_conf.get("bottom"):
            bottom = Side(**border_conf["bottom"])
        else:
            bottom = cell.border.bottom
        if border_conf.get("left"):
            left = Side(**border_conf["left"])
        else:
            left = cell.border.left
        if border_conf.get("right"):
            right = Side(**border_conf["right"])
        else:
            right = cell.border.right
        cell.border = Border(top=top, bottom=bottom, left=left, right=right)

    @staticmethod
    def _set_headers_span(col_headers):
        child_span = 1
        child_count = 1
        for header in reversed(col_headers):
            header.set_span(child_count * child_span)
            child_count = header.count
            child_span = header.span

    # 数字からエクセル列アルファベットへの変換関数（1→A、27→AAみたいな変換）
    @staticmethod
    def _to_alpha(num):
        h=int((num-1-26)/(26*26))
        i=int((num-1-(h*26*26))/26)
        j=int(num-(i*26)-(h*26*26))
        alpha=''
        for k in h,i,j:
            if k!=0:
                alpha+=chr(k+64)
        return alpha

class ColumnHeader:
    def __init__(self, header_conf, header_data):
        self.count = len(header_data)
        ds_conf = header_conf["source"]
        sorted_header_data = sorted(header_data, key=lambda x: x[ds_conf["order"]])
        keys = [x[ds_conf["data"]] for x in sorted_header_data]
        self.indexes = {}
        for k, v in zip(keys, range(len(header_data))):
            # 0始まりの連番を振り直す
            self.indexes[k] = v
        self.merge = header_conf.get("merge")

        h_r_span = header_conf.get("row_span")
        self.row_span = h_r_span if h_r_span is not None else 1
        h_r_offset = header_conf.get("row_offset")
        self.row_offset = h_r_offset if h_r_offset is not None else 0

        self.header_title_conf = header_conf.get("header_title")
        self.style_conf = header_conf.get("style")
        self.last_col_border_conf = header_conf.get("last_col_border")

    def set_span(self, span):
        self.span = span

    def set_parent_count(self, span):
        self.span = span


class RowHeader:
    def __init__(self, header_conf, header_data):
        self.count = len(header_data)
        ds_conf = header_conf["source"]
        sorted_header_data = sorted(header_data, key=lambda x: x[ds_conf["order"]])
        keys = [x[ds_conf["data"]] for x in sorted_header_data]
        self.indexes = {}
        for k, v in zip(keys, range(len(header_data))):
            # 0始まりの連番を振り直す
            self.indexes[k] = v
        self.style_conf = header_conf.get("style")

class Body:
    def __init__(self, body_conf, body_data, row_header, col_headers):
        self.row_header = row_header
        self.style_conf = body_conf.get("style")
        self.last_row_border_conf = body_conf.get("last_row_border")

        ds_conf = body_conf["source"]
        group_conf = ds_conf["group"]
        self._row_header_column_name = group_conf["row_header"]["column_name"]
        c_h_confs = sorted(group_conf["col_headers"], key=lambda x: x["header_index"])
        self._col_header_column_names = [x["column_name"] for x in c_h_confs]

        self.has_null_error = False
        self.items = []
        for item in body_data:
            if item[self._row_header_column_name] in row_header.indexes:
                row_i = row_header.indexes[item[self._row_header_column_name]] + 1
            else:
                row_i = len(row_header.indexes) + 1
                self.has_null_error = True
                logger.error("No master data: " + item[self._row_header_column_name])
            col_i = self._calc_col_index(self._col_header_column_names, col_headers, item)
            self.items.append(BodyItem(row_i, col_i, item[ds_conf["data"]]))

    @staticmethod
    def _calc_col_index(col_header_column_names, col_headers, item):
        c_index = 1
        for column_name, col_header in zip(col_header_column_names, col_headers):
            c_index += int(col_header.indexes[item[column_name]]) * col_header.span
        return c_index

class BodyItem:
    def __init__(self, row_index, col_index, value):
        self.row_index = row_index
        self.col_index = col_index
        self.value = value

